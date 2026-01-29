#!/usr/bin/env python
"""
Test script to verify the export header change from "Date of Birth" to "dob"
"""
import os
import sys
import django

# Setup Django environment
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from registration.admin import _build_export_workbook
from registration.models import CandidateProfile
from openpyxl import load_workbook
from io import BytesIO

def test_export_header():
    print("Testing export header change...")
    
    # Get a small queryset for testing
    queryset = CandidateProfile.objects.all()[:1]  # Just one candidate for testing
    
    if not queryset.exists():
        print("❌ No candidates found in database. Please add some test data first.")
        return False
    
    try:
        # Generate the export workbook
        xlsx_bytes = _build_export_workbook(queryset)
        
        # Load the workbook to check headers
        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb.active
        
        # Get the headers (first row)
        headers = []
        for cell in ws[1]:  # First row
            if cell.value:
                headers.append(cell.value)
        
        print(f"Found headers: {headers}")
        
        # Check if "dob" is present and "Date of Birth" is not
        if "dob" in headers:
            print("✅ SUCCESS: 'dob' header found in export")
        else:
            print("❌ FAILED: 'dob' header not found in export")
            return False
            
        if "Date of Birth" in headers:
            print("❌ FAILED: 'Date of Birth' header still present (should be changed to 'dob')")
            return False
        else:
            print("✅ SUCCESS: 'Date of Birth' header successfully changed")
        
        # Find the position of dob header
        try:
            dob_index = headers.index("dob")
            print(f"✅ 'dob' header found at position {dob_index + 1}")
        except ValueError:
            print("❌ Could not find 'dob' in headers list")
            return False
        
        print("\n🎉 Export header change test PASSED!")
        print("The PO export .dat files will now show 'dob' instead of 'Date of Birth'")
        return True
        
    except Exception as e:
        print(f"❌ Error during test: {str(e)}")
        return False

if __name__ == "__main__":
    success = test_export_header()
    sys.exit(0 if success else 1)